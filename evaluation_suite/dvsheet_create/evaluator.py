"""
DVSheet Create task evaluator using a text-mode VLM judge.

Evaluation approach:
1) Functionality 30%: chart existence, non-overlap ratio, and dynamic series coverage.
2) Visual quality 70%: image, text summary, user instruction, and rubric.

Model calls read config.model_config, and prompt construction lives in prompt.py.
"""

from __future__ import annotations

import argparse
import base64
import re
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Optional, Tuple, List, Any, Dict

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import range_boundaries
from openpyxl.chart.data_source import NumRef, StrRef

from evaluation_suite.dvsheet_create.config import model_config
from evaluation_suite.dvsheet_create.prompt import RUBRIC_PROMPT


# ---------------------------------------------------------------------------
# Core evaluation
# ---------------------------------------------------------------------------

@dataclass
class EvalResult:
    s_func: float
    s_vis: float
    score: float
    debug: str
    model_raw: str = ""
    prompt_used: str = ""
    vlm_total_raw: float = 0.0
    vlm_total_norm: float = 0.0
    vlm_dims_raw: Dict[str, float] = None
    vlm_dims_norm: Dict[str, float] = None
    vlm_max_scores: Dict[str, float] = None
    spatial_score: float = 0.0
    dynamic_score: float = 0.0
    table_score: float = 0.0


def evaluate_task(
    sheet: Worksheet,
    query_text: str,
    rubric_text: str,
    max_scores: Dict[str, float],
    task_type: str,
    workbook_path: Optional[Path] = None,
    chart_img: Optional[Path] = None,
    model_name: str = "gemini-2.5-flash",
    gold_workbook: Optional[Path] = None,
    combine_mode: str = "product",
    vis_weight: float = 0.7,
    table_weight: float = 0.3,
) -> EvalResult:
    """
    Evaluate a DVSheet-Create submission on a single worksheet in text mode.

    :param sheet: openpyxl worksheet containing the chart to grade.
    :param query_text: Original user instruction.
    :param rubric_text: Visual scoring rubric text.
    :param task_type: free-form string for logging (e.g., "bubble").
    :param workbook_path: Original Excel path used for table extraction.
    :param model_name: Model config name from config.model_config.
    """
    debug_lines = []

    # --- Gate checks ---
    chart = first_chart(sheet)
    if chart is None:
        debug_lines.append("No chart found; score = 0.")
        return EvalResult(0.0, 0.0, 0.0, "\n".join(debug_lines), model_raw="")

    # Spatial gate: chart should not overlap data area beyond tolerance
    score_spatial, overlap_area, data_area = check_overlap(chart, sheet)
    if score_spatial <= 0:
        debug_lines.append(f"Overlap gate failed (overlap={overlap_area}, data_area={data_area}); score=0.")
        return EvalResult(0.0, 0.0, 0.0, "\n".join(debug_lines), model_raw="")
    debug_lines.append(
        f"Spatial gate passed (non-overlap ratio): {score_spatial:.3f} (overlap={overlap_area}, data_area={data_area})"
    )

    # Dynamic gate: must reference cell ranges (not hard-coded)
    score_dynamic, dyn_count, total_series = blind_injection_test(chart)
    if score_dynamic <= 0:
        debug_lines.append(f"Dynamic gate failed ({dyn_count}/{total_series}); score=0.")
        return EvalResult(0.0, 0.0, 0.0, "\n".join(debug_lines), model_raw="")
    debug_lines.append(f"Dynamic gate passed: ratio={score_dynamic:.3f} ({dyn_count}/{total_series})")

    # --- Visual scoring only (rubric via VLM) ---
    chart_text = chart_to_text(chart, sheet)
    if chart_img is None:
        debug_lines.append("No chart image provided; visual score = 0.")
        s_vis = 0.0
        vlm_response: Dict[str, Any] = {
            "total_raw": 0.0,
            "total_norm": 0.0,
            "dims_raw": {},
            "dims_norm": {},
            "raw": "missing chart image",
            "prompt": "",
        }
    else:
        vlm_response = vlm_judge_mm(
            query_text=query_text,
            rubric_text=rubric_text,
            max_scores=max_scores,
            chart_text=chart_text,
            chart_img=chart_img,
            model_name=model_name,
        )
        s_vis = vlm_response.get("total_norm", 0.0)
        debug_lines.append(
            f"VLM total raw: {vlm_response.get('total_raw', 0.0)}, total_norm: {vlm_response.get('total_norm', 0.0)}"
        )

    # Table coverage score (0-1); if unavailable, defaults to 1.0 to avoid unfair zeroing.
    table_score = _table_coverage(sheet, workbook_path, gold_workbook, debug_lines)

    if combine_mode == "weighted":
        total = vis_weight + table_weight
        if total <= 0:
            vis_w = 0.7
            tab_w = 0.3
        else:
            vis_w = vis_weight / total
            tab_w = table_weight / total
        final_score = vis_w * s_vis + tab_w * table_score
        debug_lines.append(
            f"combine=weighted vis_w={vis_w:.3f} table_w={tab_w:.3f} -> score={final_score:.3f} (vis={s_vis:.3f}, table={table_score:.3f})"
        )
    else:
        # Default: multiplicative to enforce both chart correctness and table fidelity
        final_score = s_vis * table_score
        debug_lines.append(f"combine=product visual_only={s_vis:.3f}, table_score={table_score:.3f}")

    return EvalResult(
        s_func=0.0,
        s_vis=s_vis,
        score=final_score,
        debug="\n".join(debug_lines),
        model_raw=str(vlm_response.get("raw", "")),
        prompt_used=str(vlm_response.get("prompt", "")),
        vlm_total_raw=float(vlm_response.get("total_raw", 0.0)),
        vlm_total_norm=float(vlm_response.get("total_norm", 0.0)),
        vlm_dims_raw=vlm_response.get("dims_raw", {}),
        vlm_dims_norm=vlm_response.get("dims_norm", {}),
        vlm_max_scores=vlm_response.get("max_scores", {}),
        spatial_score=score_spatial,
        dynamic_score=score_dynamic,
        table_score=table_score,
    )


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def first_chart(sheet: Worksheet):
    charts = getattr(sheet, "_charts", [])
    return charts[0] if charts else None


def used_range_bbox(sheet: Worksheet) -> Tuple[int, int, int, int]:
    """
    Return bounding box of non-empty cells (0-based inclusive): (min_col, min_row, max_col, max_row).
    - Based only on the current sheet.
    - Ignore None values and blank strings after stripping whitespace.
    - If there are no non-empty cells, return (0,0,-1,-1); downstream area is 0.
    """
    min_c = min_r = None
    max_c = max_r = None

    for r_idx, row in enumerate(sheet.iter_rows(values_only=False), start=1):
        for c_idx, cell in enumerate(row, start=1):
            val = cell.value
            if val is None:
                continue
            if isinstance(val, str) and val.strip() == "":
                continue
            if min_c is None:
                min_c = max_c = c_idx
                min_r = max_r = r_idx
            else:
                min_c = min(min_c, c_idx)
                max_c = max(max_c, c_idx)
                min_r = min(min_r, r_idx)
                max_r = max(max_r, r_idx)

    if min_c is None:
        return 0, 0, -1, -1
    return min_c - 1, min_r - 1, max_c - 1, max_r - 1


def chart_bbox(chart, sheet: Worksheet, default_cols: int = 6, default_rows: int = 12) -> Tuple[int, int, int, int]:
    """
    Extract chart bounding box in 0-based inclusive cell indices.
    - If _to exists, use the two-point anchor directly.
    - If only a one-point anchor exists, use a fixed default size to avoid overly strict overlap penalties.
    """
    anchor = chart.anchor
    frm = getattr(anchor, "_from", None)
    to = getattr(anchor, "_to", None)
    if frm and to:
        return frm.col, frm.row, to.col, to.row
    if frm:
        return (
            frm.col,
            frm.row,
            frm.col + max(0, default_cols - 1),
            frm.row + max(0, default_rows - 1),
        )
    raise ValueError("Chart anchor missing _from/_to")


def check_overlap(chart, sheet: Worksheet) -> Tuple[float, int, int]:
    """
    Compute the non-overlap ratio: 1 - (overlap_area / data_area).
    The denominator is the active data rectangle and the numerator is chart/data overlap.
    Return (score, overlap_area, data_area).
    """
    c_min_c, c_min_r, c_max_c, c_max_r = chart_bbox(chart, sheet)
    d_min_c, d_min_r, d_max_c, d_max_r = used_range_bbox(sheet)

    data_width = max(0, d_max_c - d_min_c + 1)
    data_height = max(0, d_max_r - d_min_r + 1)
    data_area = data_width * data_height
    if data_area == 0:
        return 1.0, 0, 0

    overlap_w = max(0, min(c_max_c, d_max_c) - max(c_min_c, d_min_c) + 1)
    overlap_h = max(0, min(c_max_r, d_max_r) - max(c_min_r, d_min_r) + 1)
    overlap_area = overlap_w * overlap_h

    score = 1.0 - (overlap_area / data_area)
    return max(0.0, min(1.0, score)), overlap_area, data_area


def _datasrc_has_ref(ds) -> bool:
    if ds is None:
        return False
    # numRef / strRef mean the chart is bound to cells, not literal values.
    nr = getattr(ds, "numRef", None)
    sr = getattr(ds, "strRef", None)
    return (nr is not None and bool(getattr(nr, "f", None))) or (sr is not None and bool(getattr(sr, "f", None)))


def _series_is_dynamic(ser) -> bool:
    """
    Checks whether a series uses cell references (numRef/strRef) for all data parts present.
    """
    fields = [
        getattr(ser, "val", None),
        getattr(ser, "xVal", None),
        getattr(ser, "yVal", None),
        getattr(ser, "bubbleSize", None),
    ]
    present = [ds for ds in fields if ds is not None]
    if not present:
        return False
    return all(_datasrc_has_ref(ds) for ds in present)


def blind_injection_test(chart) -> Tuple[float, int, int]:
    """
    Dynamic series coverage: dynamic series count / total series count.
    """
    series = getattr(chart, "series", [])
    total = len(series)
    if total == 0:
        return 0.0, 0, 0
    dyn_count = sum(1 for ser in series if _series_is_dynamic(ser))
    ratio = dyn_count / total
    return ratio, dyn_count, total


# ---------------------------------------------------------------------------
# Textual extraction for visual scoring
# ---------------------------------------------------------------------------

def sheet_to_markdown(sheet: Worksheet, max_rows: int = 200, max_cols: int = 20) -> str:
    """
    Convert used range to a lightweight markdown table for LLM consumption.
    Caps rows/cols to avoid huge payloads.
    """
    min_c, min_r, max_c, max_r = range_boundaries(sheet.calculate_dimension())
    max_r = min(max_r, min_r + max_rows - 1)
    max_c = min(max_c, min_c + max_cols - 1)

    rows: List[List[str]] = []
    for r in range(min_r, max_r + 1):
        row_vals = []
        for c in range(min_c, max_c + 1):
            val = sheet.cell(row=r, column=c).value
            row_vals.append("" if val is None else str(val))
        rows.append(row_vals)

    if not rows:
        return ""

    header = rows[0]
    md = ["| " + " | ".join(header) + " |", "|" + "|".join([" --- "] * len(header)) + "|"]
    for row in rows[1:]:
        md.append("| " + " | ".join(row) + " |")
    return "\n".join(md)


def _extract_ref_values(sheet: Worksheet, ref: Any) -> List[Any]:
    if ref is None:
        return []
    rng = None
    if isinstance(ref, NumRef):
        rng = ref.f
    elif isinstance(ref, StrRef):
        rng = ref.f
    elif hasattr(ref, "numRef"):
        rng = getattr(ref.numRef, "f", None)
    elif hasattr(ref, "strRef"):
        rng = getattr(ref.strRef, "f", None)
    if not rng:
        return []
    # Guard against malformed refs like "$G$4,Sheet1"
    rng_core = rng.split("!")[1] if "!" in rng else rng
    if "," in rng_core:
        rng_core = rng_core.split(",")[0]
    try:
        min_c, min_r, max_c, max_r = range_boundaries(rng_core)
    except Exception:
        return []
    values = []
    for r in range(min_r, max_r + 1):
        for c in range(min_c, max_c + 1):
            values.append(sheet.cell(row=r, column=c).value)
    return values


def _normalize_header(h: Any) -> str:
    if h is None:
        return ""
    txt = str(h).lower().strip()
    import re
    txt = re.sub(r"\(.*?\)", "", txt)  # remove units
    # split camelCase / spaces / underscores / hyphens
    tokens = re.split(r"[\\s_\\-]+", txt)
    split_camel = []
    for t in tokens:
        split_camel.extend(re.findall(r"[a-z]+", re.sub(r"[^a-z]", "", t)) or [t])
    return "".join(split_camel)


def _table_coverage(cand_sheet: Worksheet, cand_wb_path: Optional[Path], gold_wb_path: Optional[Path], debug_lines: list[str]) -> float:
    # If no gold table provided, fallback to 1.0 (do not penalize)
    if not gold_wb_path or not cand_wb_path or not gold_wb_path.exists() or not cand_wb_path.exists():
        return 1.0
    try:
        cand_wb = load_workbook(cand_wb_path)
        gold_wb = load_workbook(gold_wb_path)
    except Exception as exc:
        debug_lines.append(f"Table load failed: {exc}")
        return 1.0

    def pick_sheet(wb):
        for ws in wb.worksheets:
            if getattr(ws, "_charts", []):
                return ws
        return wb.active

    # Prefer the sheet passed in (typically the chart sheet, often named "result"); otherwise pick by chart.
    cand_sheet = cand_sheet or pick_sheet(cand_wb)
    gold_sheet = pick_sheet(gold_wb)

    def headers_and_rows(ws: Worksheet):
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return [], []
        headers_raw = rows[0]
        headers = [_normalize_header(h) for h in headers_raw]
        data_rows = rows[1:]
        return headers, data_rows

    cand_headers, cand_rows = headers_and_rows(cand_sheet)
    gold_headers, gold_rows = headers_and_rows(gold_sheet)
    if not gold_rows or not gold_headers:
        return 1.0

    # map gold header -> cand header by exact normalized name; if not found, fallback by type
    mapping = {}
    used = set()
    for g in gold_headers:
        if g in cand_headers and g not in used:
            mapping[g] = g
            used.add(g)

    def infer_col_type(rows, idx):
        import datetime
        nums = dates = total = 0
        for r in rows:
            if idx >= len(r):
                continue
            v = r[idx]
            if v is None or v == "":
                continue
            total += 1
            if isinstance(v, (int, float)):
                nums += 1
            elif isinstance(v, (datetime.date, datetime.datetime)):
                dates += 1
        if total == 0:
            return "cat"
        if nums / total > 0.7:
            return "num"
        if dates / total > 0.3:
            return "date"
        return "cat"

    gold_types = [infer_col_type(gold_rows, i) for i, _ in enumerate(gold_headers)]
    cand_types = [infer_col_type(cand_rows, i) for i, _ in enumerate(cand_headers)]

    for gi, g in enumerate(gold_headers):
        if g in mapping:
            continue
        best_idx = None
        for ci, c in enumerate(cand_headers):
            if c in used:
                continue
            if gold_types[gi] != cand_types[ci]:
                continue
            best_idx = ci
            break
        if best_idx is not None:
            mapping[g] = cand_headers[best_idx]
            used.add(cand_headers[best_idx])

    if not mapping:
        return 0.0

    # build column value multisets (ignoring row order) excluding empty cells
    from collections import Counter

    def col_counter(rows, headers):
        cols = {h: Counter() for h in headers}
        for row in rows:
            for h, v in zip(headers, row):
                if h and v not in (None, ""):
                    cols[h][str(v).strip()] += 1
        return cols

    gold_counts = col_counter(gold_rows, gold_headers)
    cand_counts = col_counter(cand_rows, cand_headers)

    matched_cells = 0
    total_cells = 0
    for g, c in mapping.items():
        g_cnt = gold_counts.get(g) or Counter()
        c_cnt = cand_counts.get(c) or Counter()
        cand_total = sum(c_cnt.values())
        total_cells += cand_total
        for val, cnum in c_cnt.items():
            gnum = g_cnt.get(val, 0)
            matched_cells += min(gnum, cnum)

    coverage = matched_cells / total_cells if total_cells > 0 else 0.0
    debug_lines.append(
        f"Table coverage={coverage:.3f} (matched={matched_cells}, total_pred_cells={total_cells}, mapped_cols={len(mapping)})"
    )
    return coverage


def chart_to_text(chart, sheet: Worksheet) -> str:
    """
    Summarize chart series into plain text for LLM input.
    """
    lines = []
    for idx, ser in enumerate(getattr(chart, "series", []), start=1):
        title = getattr(ser, "title", None)
        if title and hasattr(title, "v"):
            title_val = title.v
        else:
            title_val = str(title) if title else f"series_{idx}"

        x_vals = _extract_ref_values(sheet, getattr(ser, "xVal", None) or getattr(ser, "cat", None))
        y_vals = _extract_ref_values(sheet, getattr(ser, "yVal", None) or getattr(ser, "val", None))
        sizes = _extract_ref_values(sheet, getattr(ser, "bubbleSize", None))

        lines.append(
            f"Series {idx} ({title_val}): X={x_vals}, Y={y_vals}" + (f", Size={sizes}" if sizes else "")
        )
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Visual scoring (text to VLM)
# ---------------------------------------------------------------------------

def _parse_score_from_text(content: str) -> float:
    match = re.search(r"Total_Score\s*[:=]?\s*([+-]?[0-9]+(?:\.[0-9]+)?)", content)
    if match:
        return float(match.group(1))
    match = re.search(r"([+-]?[0-9]+(?:\.[0-9]+)?)\s*/?\s*10", content)
    if match:
        return float(match.group(1))
    match = re.search(r"score\s*[:=]\s*([+-]?[0-9]+(?:\.[0-9]+)?)", content, re.IGNORECASE)
    if match:
        return float(match.group(1))
    return 0.0


def _extract_json(content: str) -> Optional[dict]:
    import json

    # Try direct JSON
    try:
        return json.loads(content)
    except Exception:
        pass
    # Try to find first JSON object in text/code block
    import re

    match = re.search(r"\{.*\}", content, re.DOTALL)
    if match:
        try:
            return json.loads(match.group(0))
        except Exception:
            return None
    return None


def _normalize(value: float, min_v: float, max_v: float) -> float:
    if max_v == min_v:
        return 0.0
    norm = (value - min_v) / (max_v - min_v)
    return max(0.0, min(1.0, norm))


def _parse_vlm_json(content: str, max_scores: Dict[str, float]) -> Dict[str, Any]:
    """
    Parse JSON returned by the model.
    - Dimensions: Fidelity / Logic / Aesthetics, with numbered dimensions mapped by order.
    - Total score: total or Total_Score.
    Normalize each dimension by raw/max and compute a 4:3:3 weighted percentile total.
    """
    default_max = {
        "Fidelity": max_scores.get("Fidelity", 8),
        "Logic": max_scores.get("Logic", 5),
        "Aesthetics": max_scores.get("Aesthetics", 7),
        "Total": max_scores.get("Total", 20),
    }
    weights = {"Fidelity": 4, "Logic": 3, "Aesthetics": 3}

    parsed = _extract_json(content)
    dims_raw: Dict[str, float] = {}
    dims_norm: Dict[str, float] = {}
    total_raw = None

    def assign_dim(name: str, value: Any):
        try:
            raw = float(value)
        except Exception:
            raw = 0.0
        dims_raw[name] = raw
        max_v = default_max.get(name, 1.0)
        dims_norm[name] = _normalize(raw, 0.0, max_v)

    if parsed and isinstance(parsed, dict):
        scores = parsed.get("scores") or parsed.get("score") or parsed
        if isinstance(scores, dict):
            for key in ["Fidelity", "Logic", "Aesthetics"]:
                if key in scores:
                    val = scores[key]
                    if isinstance(val, dict) and "score" in val:
                        assign_dim(key, val.get("score", 0.0))
                    else:
                        assign_dim(key, val)
            if not dims_raw:
                for idx, key in enumerate(["Dimension_1", "Dimension_2", "Dimension_3"], start=1):
                    if key in scores:
                        dim_name = ["Fidelity", "Logic", "Aesthetics"][idx - 1]
                        val_dict = scores[key] if isinstance(scores[key], dict) else {}
                        if isinstance(val_dict, dict) and "Total_Score" in val_dict:
                            assign_dim(dim_name, val_dict["Total_Score"])
                        else:
                            assign_dim(dim_name, val_dict if isinstance(val_dict, (int, float)) else 0.0)
            tot_val = scores.get("total")
            if tot_val is None and "Total_Score" in scores:
                tot_val = scores.get("Total_Score")
            if isinstance(tot_val, dict) and "score" in tot_val:
                total_raw = float(tot_val.get("score", 0.0))
            elif tot_val is not None:
                try:
                    total_raw = float(tot_val)
                except Exception:
                    total_raw = None
        if total_raw is None and isinstance(parsed, dict) and "Total_Score" in parsed:
            try:
                total_raw = float(parsed["Total_Score"])
            except Exception:
                total_raw = None

    if total_raw is None and dims_raw:
        total_raw = sum(dims_raw.values())

    if dims_norm:
        weighted = (
            dims_norm.get("Fidelity", 0.0) * weights["Fidelity"]
            + dims_norm.get("Logic", 0.0) * weights["Logic"]
            + dims_norm.get("Aesthetics", 0.0) * weights["Aesthetics"]
        )
        total_norm = weighted / sum(weights.values())
    else:
        total_raw = total_raw if total_raw is not None else 0.0
        total_norm = _normalize(total_raw, 0.0, default_max["Total"])

    return {
        "total_raw": total_raw if total_raw is not None else 0.0,
        "total_norm": total_norm,
        "dims_raw": dims_raw,
        "dims_norm": dims_norm,
        "max_scores": default_max,
    }


def _make_client(cfg):
    from openai import AzureOpenAI, OpenAI

    api_key = cfg["api_key"]
    base_url = cfg["base_url"]
    api_version = cfg.get("api_version")
    headers = cfg.get("headers") or {}

    if api_version:
        return AzureOpenAI(azure_endpoint=base_url, api_key=api_key, api_version=api_version, default_headers=headers)
    return OpenAI(base_url=base_url, api_key=api_key, default_headers=headers)


def _file_to_data_url(path: Path) -> str:
    b64 = base64.b64encode(path.read_bytes()).decode("ascii")
    return f"data:image/png;base64,{b64}"


def _parse_rubric_scores(content: str, max_scores: Dict[str, float]) -> Dict[str, Any]:
    """
    Parse rubric JSON:
    - Expect per-dimension dicts with a "subtotal" and nested standards.
    - Total_Score is preferred; otherwise sum of subtotals.
    """
    dim_scores: Dict[str, float] = {}
    dim_percents: Dict[str, float] = {}
    total_raw = 0.0
    parsed = _extract_json(content)
    if parsed and isinstance(parsed, dict):
        data = parsed
        for dim_key, dim_val in data.items():
            if not isinstance(dim_val, dict):
                continue
            subtotal = dim_val.get("subtotal")
            if subtotal is None:
                # fallback: sum all child scores under this dimension
                subtotal_sum = 0.0
                for std_key, std_val in dim_val.items():
                    if not isinstance(std_val, dict):
                        continue
                    sc = std_val.get("score")
                    if sc is None:
                        continue
                    try:
                        subtotal_sum += float(sc)
                    except Exception:
                        continue
                subtotal = subtotal_sum
            if subtotal is not None:
                try:
                    dim_scores[dim_key] = float(subtotal)
                except Exception:
                    pass
        if "Total_Score" in data:
            try:
                total_raw = float(data.get("Total_Score", 0.0))
            except Exception:
                total_raw = 0.0
        else:
            total_raw = sum(dim_scores.values())
    else:
        total_raw = _parse_score_from_text(content)

    if dim_scores and max_scores:
        canonical = ["Fidelity", "Logic", "Aesthetics"]
        if all(k.startswith("Dimension") for k in dim_scores.keys()) and all(k in max_scores for k in canonical):
            reordered = {}
            for idx, (k, v) in enumerate(sorted(dim_scores.items())):
                if idx < len(canonical):
                    reordered[canonical[idx]] = v
                else:
                    reordered[k] = v
            dim_scores = reordered

    total_points = max_scores.get("Total") if max_scores else None
    total_norm = max(0.0, min(1.0, total_raw / total_points)) if total_points and total_points > 0 else 0.0

    if dim_scores and max_scores:
        for dk, raw in dim_scores.items():
            max_v = max_scores.get(dk)
            if max_v:
                dim_percents[dk] = max(0.0, min(1.0, raw / max_v))

    return {
        "dims_raw": dim_scores,
        "dims_norm": dim_percents,
        "total_raw": total_raw,
        "total_norm": total_norm,
        "max_scores": max_scores or {},
        "items": {},
    }


def vlm_judge_mm(query_text: str, rubric_text: str, max_scores: Dict[str, float], chart_text: str, chart_img: Path, model_name: str):
    """
    Multimodal evaluation: text + image.
    """
    if model_name not in model_config:
        raise ValueError(f"Model config '{model_name}' not found.")
    cfg = model_config[model_name]
    client = _make_client(cfg)

    prompt = RUBRIC_PROMPT.format(
        user_query=query_text.strip(),
        rubric=rubric_text.strip(),
        chart_data=chart_text.strip(),
    )
    messages = [
        {
            "role": "user",
            "content": [
                {"type": "text", "text": prompt},
                {"type": "image_url", "image_url": {"url": _file_to_data_url(chart_img)}},
            ],
        }
    ]

    response = None
    last_err = ""
    for _ in range(3):
        try:
            response = client.chat.completions.create(
                model=cfg["model_name"],
                messages=messages,
                stream=False,
                **cfg.get("generate_kwargs", {}),
            )
            break
        except Exception as exc:  # noqa: BLE001
            last_err = str(exc)
            time.sleep(0.5)

    content = response.choices[0].message.content if response and response.choices else ""
    if not content and last_err:
        content = f"[error] {last_err}"
    parsed = _parse_rubric_scores(content or "", max_scores)
    return {
        "total_raw": parsed["total_raw"],
        "total_norm": parsed["total_norm"],
        "dims_raw": parsed["dims_raw"],
        "dims_norm": parsed["dims_norm"],
        "raw": content,
        "prompt": prompt,
        "max_scores": parsed.get("max_scores", {}),
        "items": parsed.get("items", {}),
    }


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Evaluate DVSheet-Create tasks (text-based VLM).")
    parser.add_argument("workbook", type=Path, help="Path to the Excel workbook.")
    parser.add_argument("--sheet", help="Worksheet name; defaults to first sheet with a chart.")
    parser.add_argument("--query-file", type=Path, required=True, help="Path to query text (e.g., gold/.../query.md).")
    parser.add_argument("--rubric-file", type=Path, required=True, help="Path to rubric text (e.g., gold/.../rubric.md).")
    parser.add_argument("--chart-img", type=Path, required=False, help="PNG of the candidate chart (optional, but needed for visual score).")
    parser.add_argument("--model", default="gemini-2.5-flash", help="Model name key from config.model_config.")
    args = parser.parse_args()

    query_text = args.query_file.read_text(encoding="utf-8")
    rubric_text = args.rubric_file.read_text(encoding="utf-8")

    wb = load_workbook(args.workbook)
    if args.sheet:
        ws = wb[args.sheet]
    else:
        ws = None
        for candidate in wb.worksheets:
            if getattr(candidate, "_charts", []):
                ws = candidate
                break
        if ws is None:
            raise ValueError("No chart found in workbook; specify --sheet explicitly.")

    res = evaluate_task(
        sheet=ws,
        query_text=query_text,
        rubric_text=rubric_text,
        task_type="auto",
        workbook_path=args.workbook,
        chart_img=args.chart_img,
        model_name=args.model,
    )

    print(f"s_func={res.s_func:.3f}, s_vis={res.s_vis:.3f}, final={res.score:.3f}")
    print("--- debug ---")
    print(res.debug)


if __name__ == "__main__":
    main()
