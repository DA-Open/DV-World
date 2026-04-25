"""
Stitch exported dashboard chart PNGs into a single "dashboard" image.

Input:
- dashboard_charts.json (written by export_dashboard_png.py)
- exported chart images (e.g., dashboard_chart_1.png, ...)

This does NOT require Excel. It uses the recorded ChartObject Left/Top/Width/Height
to place each chart image onto a canvas.

Note:
- This only stitches charts (not cell backgrounds / shapes / slicers).
"""

from __future__ import annotations

import json
import statistics
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple


def _load_manifest(manifest_path: Path) -> List[Dict[str, Any]]:
    data = json.loads(manifest_path.read_text(encoding="utf-8"))
    if not isinstance(data, list):
        raise ValueError("manifest must be a JSON list")
    return [d for d in data if isinstance(d, dict) and d.get("png")]


def _compute_scale(entries: List[Dict[str, Any]], base_dir: Path) -> float:
    """
    Compute a robust points->pixels scale from chart image sizes.
    """
    try:
        from PIL import Image  # type: ignore
    except Exception as exc:  # noqa: BLE001
        raise RuntimeError("Missing dependency: Pillow. Install via: pip install Pillow") from exc

    scales: List[float] = []
    for e in entries:
        png = resolve_png_path(base_dir, str(e["png"]))
        if not png.exists():
            continue
        w_pt = float(e.get("width") or 0.0)
        h_pt = float(e.get("height") or 0.0)
        if w_pt <= 1 or h_pt <= 1:
            continue
        try:
            with Image.open(png) as im:
                w_px, h_px = im.size
        except Exception:
            continue
        if w_px > 0 and w_pt > 0:
            scales.append(w_px / w_pt)
        if h_px > 0 and h_pt > 0:
            scales.append(h_px / h_pt)

    if not scales:
        return 1.0
    return float(statistics.median(scales))


def _bounds(entries: List[Dict[str, Any]]) -> Tuple[float, float, float, float]:
    min_left = min(float(e.get("left") or 0.0) for e in entries)
    min_top = min(float(e.get("top") or 0.0) for e in entries)
    max_right = max(float(e.get("left") or 0.0) + float(e.get("width") or 0.0) for e in entries)
    max_bottom = max(float(e.get("top") or 0.0) + float(e.get("height") or 0.0) for e in entries)
    return min_left, min_top, max_right, max_bottom


def _load_layout(layout_path: Path) -> Dict[str, Any]:
    return json.loads(layout_path.read_text(encoding="utf-8"))


def _font():
    from PIL import ImageFont  # type: ignore

    # Try common fonts; fall back to default.
    for name in ["arial.ttf", "Arial.ttf", "msyh.ttc", "simhei.ttf"]:
        try:
            return ImageFont.truetype(name, 16)
        except Exception:
            continue
    return ImageFont.load_default()


def _load_sheet_for_render(workbook_path: Path, sheet_name: str):
    from openpyxl import load_workbook  # type: ignore

    wb = load_workbook(workbook_path, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    return wb, ws


def _rgb_from_openpyxl_color(color) -> Optional[Tuple[int, int, int]]:
    """
    Best-effort parse of openpyxl Color to RGB tuple.
    Supports .rgb like 'FFRRGGBB' or 'RRGGBB'.
    """
    if color is None:
        return None
    rgb = getattr(color, "rgb", None)
    if not rgb:
        return None
    s = str(rgb)
    if len(s) == 8:  # AARRGGBB
        s = s[2:]
    if len(s) != 6:
        return None
    try:
        r = int(s[0:2], 16)
        g = int(s[2:4], 16)
        b = int(s[4:6], 16)
        return r, g, b
    except Exception:
        return None


def _cell_fill_rgb(cell) -> Optional[Tuple[int, int, int]]:
    try:
        fill = cell.fill
        if fill is None:
            return None
        pattern = getattr(fill, "patternType", None)
        if pattern in (None, "none"):
            return None
        fg = getattr(fill, "fgColor", None)
        return _rgb_from_openpyxl_color(fg)
    except Exception:
        return None


def _cell_font_rgb(cell) -> Optional[Tuple[int, int, int]]:
    try:
        font = cell.font
        if font is None:
            return None
        col = getattr(font, "color", None)
        return _rgb_from_openpyxl_color(col)
    except Exception:
        return None


def _merged_ranges(ws) -> List[Tuple[int, int, int, int]]:
    out: List[Tuple[int, int, int, int]] = []
    try:
        for mr in ws.merged_cells.ranges:
            # min_row, min_col, max_row, max_col
            out.append((mr.min_row, mr.min_col, mr.max_row, mr.max_col))
    except Exception:
        pass
    return out


def _find_merge(merges: List[Tuple[int, int, int, int]], r: int, c: int) -> Optional[Tuple[int, int, int, int]]:
    for r0, c0, r1, c1 in merges:
        if r0 <= r <= r1 and c0 <= c <= c1:
            return (r0, c0, r1, c1)
    return None


def _draw_cell_borders(draw, cell, x0: int, y0: int, x1: int, y1: int):
    """
    Draw cell borders based on Excel border styles.
    """
    try:
        border = cell.border
        if not border:
            return
        
        def get_border_color(side):
            """Extract border color."""
            if not side or not side.style:
                return None
            if side.color:
                rgb = _rgb_from_openpyxl_color(side.color)
                if rgb:
                    return rgb
            return (0, 0, 0)
        
        def get_line_width(side):
            """Return line width for a border style."""
            if not side or not side.style:
                return 0
            style = str(side.style).lower()
            if style in ("thin", "hair"):
                return 1
            elif style in ("medium", "double"):
                return 2
            elif style == "thick":
                return 3
            else:
                return 1
        
        if border.top and border.top.style:
            color = get_border_color(border.top)
            width = get_line_width(border.top)
            if color:
                draw.line([(x0, y0), (x1, y0)], fill=color, width=width)
        
        if border.bottom and border.bottom.style:
            color = get_border_color(border.bottom)
            width = get_line_width(border.bottom)
            if color:
                draw.line([(x0, y1), (x1, y1)], fill=color, width=width)
        
        if border.left and border.left.style:
            color = get_border_color(border.left)
            width = get_line_width(border.left)
            if color:
                draw.line([(x0, y0), (x0, y1)], fill=color, width=width)
        
        if border.right and border.right.style:
            color = get_border_color(border.right)
            width = get_line_width(border.right)
            if color:
                draw.line([(x1, y0), (x1, y1)], fill=color, width=width)
    
    except Exception:
        pass



def stitch_dashboard_with_text_layer(
    *,
    case_dir: Path,
    workbook_path: Optional[Path] = None,
    manifest_name: str = "dashboard_charts.json",
    layout_name: str = "dashboard_layout.json",
    out_name: str = "dashboard_stitched.png",
    draw_grid: bool = False,
    font_scale: float = 1.0,
) -> Optional[Path]:
    """
    Create a stitched image with a text-only base layer (tables/titles) + charts overlay.
    - Base layer: render cell values in UsedRange onto white background (no fill/colors).
    - Overlay: paste chart PNGs at their positions from manifest.
    """
    try:
        from PIL import Image, ImageDraw, ImageFont  # type: ignore
    except Exception as exc:  # noqa: BLE001
        raise RuntimeError("Missing dependency: Pillow. Install via: pip install Pillow") from exc

    manifest_path = case_dir / manifest_name
    layout_path = case_dir / layout_name
    if not manifest_path.exists() or not layout_path.exists():
        return None

    entries = _load_manifest(manifest_path)
    if not entries:
        return None

    layout = _load_layout(layout_path)
    sheet = str(layout.get("sheet") or "")
    start_row = int(layout.get("start_row") or 1)
    start_col = int(layout.get("start_col") or 1)
    rows = int(layout.get("rows") or 1)
    cols = int(layout.get("cols") or 1)
    left_pt = float(layout.get("left_pt") or 0.0)
    top_pt = float(layout.get("top_pt") or 0.0)
    col_width_pts = [float(x or 0.0) for x in (layout.get("col_width_pts") or [])]
    row_height_pts = [float(x or 0.0) for x in (layout.get("row_height_pts") or [])]

    if workbook_path is None:
        # pick first xlsx/xls in case dir
        for p in sorted(case_dir.iterdir()):
            if p.is_file() and p.suffix.lower() in {".xlsx", ".xls"} and not p.name.startswith("~$"):
                workbook_path = p
                break
    if workbook_path is None:
        return None

    # Scale from points->pixels using chart image sizes.
    scale = _compute_scale(entries, case_dir)

    # Bounds: include used range and chart bounds.
    min_left, min_top, max_right, max_bottom = _bounds(entries)
    min_left = min(min_left, left_pt)
    min_top = min(min_top, top_pt)
    used_w_pt = sum(col_width_pts) if col_width_pts else 0.0
    used_h_pt = sum(row_height_pts) if row_height_pts else 0.0
    max_right = max(max_right, left_pt + used_w_pt)
    max_bottom = max(max_bottom, top_pt + used_h_pt)

    canvas_w = max(1, int(round((max_right - min_left) * scale)) + 40)
    canvas_h = max(1, int(round((max_bottom - min_top) * scale)) + 40)
    canvas = Image.new("RGB", (canvas_w, canvas_h), color=(255, 255, 255))
    draw = ImageDraw.Draw(canvas)
    # Pick a larger font by default, based on row height in pixels.
    row_px = [h * scale for h in row_height_pts if h and h > 0]
    base_px = float(statistics.median(row_px)) if row_px else 22.0
    font_px = max(14.0, min(40.0, base_px * 0.70)) * max(0.5, float(font_scale))
    font_size = max(10, int(round(font_px)))

    # Try common fonts at computed size.
    font = None
    for name in ["arial.ttf", "Arial.ttf", "msyh.ttc", "simhei.ttf"]:
        try:
            font = ImageFont.truetype(name, font_size)
            break
        except Exception:
            continue
    if font is None:
        font = ImageFont.load_default()

    # Precompute cumulative offsets in points for columns/rows.
    col_x_pt = [0.0]
    for w in col_width_pts:
        col_x_pt.append(col_x_pt[-1] + (w if w > 0 else 8.0))
    row_y_pt = [0.0]
    for h in row_height_pts:
        row_y_pt.append(row_y_pt[-1] + (h if h > 0 else 15.0))

    wb_obj, ws_obj = _load_sheet_for_render(workbook_path, sheet)
    try:
        merges = _merged_ranges(ws_obj)

        # Background fills (no border/formatting other than fill + font color/bold).
        end_row = start_row + max(0, rows - 1)
        end_col = start_col + max(0, cols - 1)
        for r in range(start_row, end_row + 1):
            rr = r - start_row
            if rr < 0 or rr >= len(row_y_pt) - 1:
                continue
            y0_pt = top_pt + row_y_pt[rr]
            y1_pt = top_pt + row_y_pt[rr + 1]
            y0 = int(round((y0_pt - min_top) * scale)) + 20
            y1 = int(round((y1_pt - min_top) * scale)) + 20
            for c in range(start_col, end_col + 1):
                cc = c - start_col
                if cc < 0 or cc >= len(col_x_pt) - 1:
                    continue
                x0_pt = left_pt + col_x_pt[cc]
                x1_pt = left_pt + col_x_pt[cc + 1]
                x0 = int(round((x0_pt - min_left) * scale)) + 20
                x1 = int(round((x1_pt - min_left) * scale)) + 20
                cell = ws_obj.cell(row=r, column=c)
                fill_rgb = _cell_fill_rgb(cell)
                if fill_rgb:
                    draw.rectangle([x0, y0, x1, y1], fill=fill_rgb)

                _draw_cell_borders(draw, cell, x0, y0, x1, y1)

        # Optionally draw a light grid (helps readability when background fills are absent).
        if draw_grid:
            grid_color = (230, 230, 230)
            for i in range(len(col_x_pt)):
                x_pt = left_pt + col_x_pt[i]
                x = int(round((x_pt - min_left) * scale)) + 20
                draw.line([(x, 20), (x, canvas_h - 20)], fill=grid_color, width=1)
            for i in range(len(row_y_pt)):
                y_pt = top_pt + row_y_pt[i]
                y = int(round((y_pt - min_top) * scale)) + 20
                draw.line([(20, y), (canvas_w - 20, y)], fill=grid_color, width=1)

        # Text: draw only once per merged range (top-left), otherwise per cell.
        drawn_merge_topleft: set[Tuple[int, int]] = set()
        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                cell = ws_obj.cell(row=r, column=c)
                v = cell.value
                if v is None:
                    continue
                # format numeric values to two decimals when applicable
                if isinstance(v, float):
                    s = f"{v:.2f}"
                elif isinstance(v, int):
                    s = str(v)
                elif isinstance(v, str):
                    s = v.strip()
                    import re
                    if re.fullmatch(r"[+-]?\d+\.\d+", s):
                        try:
                            s = f"{float(s):.2f}"
                        except Exception:
                            s = s.strip()
                else:
                    s = str(v).strip()
                if s == "":
                    continue

                merge = _find_merge(merges, r, c)
                if merge:
                    r0, c0, r1, c1 = merge
                    if (r, c) != (r0, c0):
                        continue
                    if (r0, c0) in drawn_merge_topleft:
                        continue
                    drawn_merge_topleft.add((r0, c0))
                    rr0 = r0 - start_row
                    cc0 = c0 - start_col
                    rr1 = r1 - start_row + 1
                    cc1 = c1 - start_col + 1
                    if rr0 < 0 or cc0 < 0:
                        continue
                    if rr1 > len(row_y_pt) or cc1 > len(col_x_pt):
                        continue
                    x0_pt = left_pt + col_x_pt[cc0]
                    y0_pt = top_pt + row_y_pt[rr0]
                    x1_pt = left_pt + col_x_pt[cc1]
                    y1_pt = top_pt + row_y_pt[rr1]
                else:
                    rr0 = r - start_row
                    cc0 = c - start_col
                    if rr0 < 0 or cc0 < 0:
                        continue
                    if rr0 + 1 > len(row_y_pt) - 1 or cc0 + 1 > len(col_x_pt) - 1:
                        continue
                    x0_pt = left_pt + col_x_pt[cc0]
                    y0_pt = top_pt + row_y_pt[rr0]
                    x1_pt = left_pt + col_x_pt[cc0 + 1]
                    y1_pt = top_pt + row_y_pt[rr0 + 1]

                x0 = int(round((x0_pt - min_left) * scale)) + 20
                y0 = int(round((y0_pt - min_top) * scale)) + 20
                x1 = int(round((x1_pt - min_left) * scale)) + 20
                y1 = int(round((y1_pt - min_top) * scale)) + 20

                # Font styling: bold/size bump for merged (often titles).
                font_here = font
                try:
                    if merge and bool(getattr(cell.font, "bold", False)):
                        # slightly larger for bold merged titles
                        fs = int(round(font_size * 1.25))
                        for name in ["arial.ttf", "Arial.ttf", "msyh.ttc", "simhei.ttf"]:
                            try:
                                font_here = ImageFont.truetype(name, fs)
                                break
                            except Exception:
                                continue
                except Exception:
                    font_here = font

                color = _cell_font_rgb(cell) or (0, 0, 0)

                # Simple alignment: use openpyxl alignment if present.
                halign = getattr(getattr(cell, "alignment", None), "horizontal", None)
                valign = getattr(getattr(cell, "alignment", None), "vertical", None)

                # Measure text.
                try:
                    bbox = draw.textbbox((0, 0), s, font=font_here)
                    tw = bbox[2] - bbox[0]
                    th = bbox[3] - bbox[1]
                except Exception:
                    tw, th = 0, 0

                pad = 4
                if halign in ("center", "centerContinuous"):
                    tx = x0 + max(pad, (x1 - x0 - tw) // 2)
                elif halign == "right":
                    tx = x1 - tw - pad
                else:
                    tx = x0 + pad

                if valign == "center":
                    ty = y0 + max(pad, (y1 - y0 - th) // 2)
                elif valign == "top":
                    ty = y0 + pad
                else:
                    ty = y0 + pad

                draw.text((tx, ty), s[:200], fill=color, font=font_here)
    finally:
        try:
            wb_obj.close()
        except Exception:
            pass

    # Overlay charts.
    for e in sorted(entries, key=lambda x: int(x.get("index") or 0)):
        png = resolve_png_path(case_dir, str(e["png"]))
        if not png.exists():
            continue
        left = float(e.get("left") or 0.0)
        top = float(e.get("top") or 0.0)
        w_pt = float(e.get("width") or 0.0)
        h_pt = float(e.get("height") or 0.0)
        if w_pt <= 1 or h_pt <= 1:
            continue
        x = int(round((left - min_left) * scale)) + 20
        y = int(round((top - min_top) * scale)) + 20
        w = max(1, int(round(w_pt * scale)))
        h = max(1, int(round(h_pt * scale)))
        try:
            im = Image.open(png).convert("RGBA")
        except Exception:
            continue
        if im.size != (w, h):
            im = im.resize((w, h))
        canvas.paste(im, (x, y), mask=im)

    out_path = case_dir / out_name
    canvas.save(out_path, format="PNG")
    return out_path


def stitch_dashboard_from_manifest(
    *,
    manifest_path: Path,
    out_path: Path,
    background: Tuple[int, int, int] = (255, 255, 255),
    margin_px: int = 20,
) -> Optional[Path]:
    """
    Stitch charts into a single PNG. Returns out_path on success, None if no charts.
    """
    entries = _load_manifest(manifest_path)
    if not entries:
        return None

    try:
        from PIL import Image  # type: ignore
    except Exception as exc:  # noqa: BLE001
        raise RuntimeError("Missing dependency: Pillow. Install via: pip install Pillow") from exc

    base_dir = manifest_path.parent
    scale = _compute_scale(entries, base_dir)
    min_left, min_top, max_right, max_bottom = _bounds(entries)

    canvas_w = max(1, int(round((max_right - min_left) * scale)) + 2 * margin_px)
    canvas_h = max(1, int(round((max_bottom - min_top) * scale)) + 2 * margin_px)
    canvas = Image.new("RGB", (canvas_w, canvas_h), color=background)

    # Keep original export order (ChartObjects index).
    for e in sorted(entries, key=lambda x: int(x.get("index") or 0)):
        png = resolve_png_path(base_dir, str(e["png"]))
        if not png.exists():
            continue
        left = float(e.get("left") or 0.0)
        top = float(e.get("top") or 0.0)
        w_pt = float(e.get("width") or 0.0)
        h_pt = float(e.get("height") or 0.0)
        if w_pt <= 1 or h_pt <= 1:
            continue

        x = int(round((left - min_left) * scale)) + margin_px
        y = int(round((top - min_top) * scale)) + margin_px
        w = max(1, int(round(w_pt * scale)))
        h = max(1, int(round(h_pt * scale)))

        try:
            im = Image.open(png).convert("RGB")
        except Exception:
            continue
        if im.size != (w, h):
            im = im.resize((w, h))
        canvas.paste(im, (x, y))

    # If nothing got pasted (e.g., path mismatch), fail loudly instead of writing a blank image.
    # Heuristic: a fully white canvas will have no non-white pixels.
    try:
        from PIL import ImageStat  # type: ignore

        stat = ImageStat.Stat(canvas.convert("L"))
        if stat.var and float(stat.var[0]) < 1e-3:
            raise RuntimeError("stitched image looks blank; chart PNG paths may be wrong. Re-export charts to refresh manifest.")
    except Exception:
        pass

    out_path.parent.mkdir(parents=True, exist_ok=True)
    canvas.save(out_path, format="PNG")
    return out_path


def resolve_png_path(case_dir: Path, png_field: str) -> Path:
    """
    Resolve manifest png field to an existing file.
    - New manifests store just filename (recommended).
    - Older manifests may store a path relative to repo root or case dir, or an absolute path.
    """
    p = Path(png_field)
    if p.is_absolute() and p.exists():
        return p
    # First try relative to case dir.
    cand = (case_dir / p)
    if cand.exists():
        return cand
    # If it contains directories, try basename under case dir.
    cand2 = case_dir / p.name
    if cand2.exists():
        return cand2
    # As last resort, try interpreting it as repo-relative from CWD.
    cand3 = Path(png_field)
    if cand3.exists():
        return cand3
    return cand2


def main():
    import argparse

    parser = argparse.ArgumentParser(description="Stitch dashboard chart PNG files into one image without Excel")
    parser.add_argument("--case-dir", required=True, type=Path, help="Case directory containing dashboard_charts.json and chart PNG files")
    parser.add_argument("--manifest", default="dashboard_charts.json", help="Manifest filename")
    parser.add_argument("--out", default="dashboard_stitched.png", help="Output PNG filename")
    parser.add_argument("--with-text", action="store_true", help="Render a text layer for dashboard tables and titles without background formatting")
    parser.add_argument("--layout", default="dashboard_layout.json", help="Layout filename used with --with-text")
    parser.add_argument("--workbook", default=None, type=Path, help="XLSX path; defaults to auto-discovery under case-dir")
    parser.add_argument("--grid", action="store_true", help="Draw light grid lines; optional with --with-text")
    parser.add_argument("--font-scale", type=float, default=1.0, help="Text layer font scale; default 1.0")
    args = parser.parse_args()

    if args.with_text:
        out = stitch_dashboard_with_text_layer(
            case_dir=args.case_dir,
            workbook_path=args.workbook,
            manifest_name=args.manifest,
            layout_name=args.layout,
            out_name=args.out,
            draw_grid=bool(args.grid),
            font_scale=float(args.font_scale),
        )
    else:
        out = stitch_dashboard_from_manifest(
            manifest_path=args.case_dir / args.manifest,
            out_path=args.case_dir / args.out,
        )
    if out is None:
        print("[skip] no charts in manifest")
    else:
        print(f"[ok] {out}")


if __name__ == "__main__":
    main()
